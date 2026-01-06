"""
Excel Export for Valuation Models.

Generates Excel workbooks with DCF and reverse DCF models.
All formulas are auditable and reproducible.
"""

from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from er.valuation.dcf import DCFInputs, DCFResult
from er.valuation.reverse_dcf import ReverseDCFInputs, ReverseDCFResult


@dataclass
class ValuationWorkbook:
    """Container for valuation model export."""

    ticker: str
    company_name: str
    dcf_result: DCFResult | None = None
    reverse_dcf_result: ReverseDCFResult | None = None
    sensitivity_data: dict[str, Any] | None = None

    def to_dict(self) -> dict[str, Any]:
        """Convert to dict for JSON export."""
        return {
            "ticker": self.ticker,
            "company_name": self.company_name,
            "dcf": self.dcf_result.to_dict() if self.dcf_result else None,
            "reverse_dcf": self.reverse_dcf_result.to_dict() if self.reverse_dcf_result else None,
            "sensitivity": self.sensitivity_data,
        }


class ValuationExporter:
    """Exports valuation models to various formats.

    Supports:
    - JSON export (always available)
    - CSV export for data
    - Excel export (if openpyxl available)
    """

    def __init__(self) -> None:
        """Initialize the exporter."""
        self._has_excel = self._check_excel_support()

    def _check_excel_support(self) -> bool:
        """Check if openpyxl is available for Excel export."""
        try:
            import openpyxl  # noqa: F401
            return True
        except ImportError:
            return False

    def export_json(
        self,
        workbook: ValuationWorkbook,
        output_path: str | Path,
    ) -> Path:
        """Export valuation model to JSON.

        Args:
            workbook: Valuation workbook data.
            output_path: Output file path.

        Returns:
            Path to created file.
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        with open(output_path, "w") as f:
            json.dump(workbook.to_dict(), f, indent=2)

        return output_path

    def export_csv(
        self,
        workbook: ValuationWorkbook,
        output_dir: str | Path,
    ) -> list[Path]:
        """Export valuation model to CSV files.

        Args:
            workbook: Valuation workbook data.
            output_dir: Output directory.

        Returns:
            List of created file paths.
        """
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        created_files = []

        # Export DCF projections
        if workbook.dcf_result:
            dcf_path = output_dir / f"{workbook.ticker}_dcf.csv"
            self._export_dcf_csv(workbook.dcf_result, dcf_path)
            created_files.append(dcf_path)

        # Export sensitivity analysis
        if workbook.sensitivity_data:
            sens_path = output_dir / f"{workbook.ticker}_sensitivity.csv"
            self._export_sensitivity_csv(workbook.sensitivity_data, sens_path)
            created_files.append(sens_path)

        # Export reverse DCF
        if workbook.reverse_dcf_result:
            rdcf_path = output_dir / f"{workbook.ticker}_reverse_dcf.csv"
            self._export_reverse_dcf_csv(workbook.reverse_dcf_result, rdcf_path)
            created_files.append(rdcf_path)

        return created_files

    def _export_dcf_csv(self, result: DCFResult, path: Path) -> None:
        """Export DCF result to CSV."""
        lines = [
            "DCF Model Output",
            "",
            "Metric,Value",
            f"Intrinsic Value per Share,{result.intrinsic_value_per_share:.2f}",
            f"Enterprise Value,{result.enterprise_value:.0f}",
            f"Equity Value,{result.equity_value:.0f}",
            f"PV of FCF,{result.pv_fcf:.0f}",
            f"Terminal Value,{result.terminal_value:.0f}",
            f"PV of Terminal,{result.pv_terminal:.0f}",
            f"WACC,{result.wacc:.2%}",
            f"Terminal Growth,{result.terminal_growth:.2%}",
            "",
            "Year,FCF,Discount Factor,PV",
        ]

        for i, (fcf, df) in enumerate(zip(result.fcf_projections, result.discount_factors)):
            pv = fcf * df
            lines.append(f"{i+1},{fcf:.0f},{df:.4f},{pv:.0f}")

        with open(path, "w") as f:
            f.write("\n".join(lines))

    def _export_sensitivity_csv(self, data: dict[str, Any], path: Path) -> None:
        """Export sensitivity analysis to CSV."""
        sensitivity = data.get("sensitivity", [])
        if not sensitivity:
            return

        # Get unique WACC values
        waccs = sorted(set(item[0] for item in sensitivity))
        tgs = sorted(set(item[1] for item in sensitivity))

        # Create matrix
        lines = ["Sensitivity Analysis: Intrinsic Value per Share", ""]

        # Header row
        header = "WACC \\ Terminal Growth," + ",".join(f"{tg:.1%}" for tg in tgs)
        lines.append(header)

        # Data rows
        for wacc in waccs:
            row_values = []
            for tg in tgs:
                value = next(
                    (item[2] for item in sensitivity if item[0] == wacc and item[1] == tg),
                    0
                )
                row_values.append(f"{value:.2f}")
            lines.append(f"{wacc:.1%}," + ",".join(row_values))

        with open(path, "w") as f:
            f.write("\n".join(lines))

    def _export_reverse_dcf_csv(self, result: ReverseDCFResult, path: Path) -> None:
        """Export reverse DCF result to CSV."""
        lines = [
            "Reverse DCF Model Output",
            "",
            "Metric,Value",
            f"Implied Revenue CAGR,{result.implied_revenue_cagr:.2%}",
            f"Market Cap,{result.market_cap:.0f}",
            f"Enterprise Value,{result.enterprise_value:.0f}",
            f"Implied Year 5 Revenue,{result.implied_year5_revenue:.0f}",
            f"Implied Year 5 FCF,{result.implied_year5_fcf:.0f}",
            f"Is Reasonable,{result.is_reasonable}",
            "",
            "Reasonableness Notes",
        ]

        for note in result.reasonableness_notes:
            lines.append(note)

        with open(path, "w") as f:
            f.write("\n".join(lines))

    def export_excel(
        self,
        workbook: ValuationWorkbook,
        output_path: str | Path,
    ) -> Path | None:
        """Export valuation model to Excel.

        Args:
            workbook: Valuation workbook data.
            output_path: Output file path.

        Returns:
            Path to created file, or None if Excel not supported.
        """
        if not self._has_excel:
            return None

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()

        # Summary sheet
        summary_ws = wb.active
        summary_ws.title = "Summary"
        self._create_summary_sheet(summary_ws, workbook)

        # DCF sheet
        if workbook.dcf_result:
            dcf_ws = wb.create_sheet("DCF Model")
            self._create_dcf_sheet(dcf_ws, workbook.dcf_result)

        # Sensitivity sheet
        if workbook.sensitivity_data:
            sens_ws = wb.create_sheet("Sensitivity")
            self._create_sensitivity_sheet(sens_ws, workbook.sensitivity_data)

        # Reverse DCF sheet
        if workbook.reverse_dcf_result:
            rdcf_ws = wb.create_sheet("Reverse DCF")
            self._create_reverse_dcf_sheet(rdcf_ws, workbook.reverse_dcf_result)

        wb.save(output_path)
        return output_path

    def _create_summary_sheet(self, ws: Any, workbook: ValuationWorkbook) -> None:
        """Create summary sheet in Excel."""
        from openpyxl.styles import Font, PatternFill

        ws["A1"] = "Valuation Summary"
        ws["A1"].font = Font(bold=True, size=14)

        ws["A3"] = "Ticker:"
        ws["B3"] = workbook.ticker

        ws["A4"] = "Company:"
        ws["B4"] = workbook.company_name

        if workbook.dcf_result:
            ws["A6"] = "DCF Intrinsic Value:"
            ws["B6"] = f"${workbook.dcf_result.intrinsic_value_per_share:.2f}"
            ws["B6"].font = Font(bold=True)

        if workbook.reverse_dcf_result:
            ws["A7"] = "Implied Growth Rate:"
            ws["B7"] = f"{workbook.reverse_dcf_result.implied_revenue_cagr:.1%}"

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

    def _create_dcf_sheet(self, ws: Any, result: DCFResult) -> None:
        """Create DCF model sheet in Excel."""
        from openpyxl.styles import Font

        ws["A1"] = "DCF Model"
        ws["A1"].font = Font(bold=True, size=12)

        # Key outputs
        ws["A3"] = "Intrinsic Value per Share"
        ws["B3"] = result.intrinsic_value_per_share
        ws["C3"] = "=B3"  # Formula reference for Excel

        ws["A4"] = "Enterprise Value"
        ws["B4"] = result.enterprise_value

        ws["A5"] = "Equity Value"
        ws["B5"] = result.equity_value

        ws["A6"] = "WACC"
        ws["B6"] = result.wacc

        ws["A7"] = "Terminal Growth"
        ws["B7"] = result.terminal_growth

        # FCF projections
        ws["A10"] = "Year"
        ws["B10"] = "FCF"
        ws["C10"] = "Discount Factor"
        ws["D10"] = "Present Value"

        for i, (fcf, df) in enumerate(zip(result.fcf_projections, result.discount_factors)):
            row = 11 + i
            ws[f"A{row}"] = i + 1
            ws[f"B{row}"] = fcf
            ws[f"C{row}"] = df
            ws[f"D{row}"] = fcf * df

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15

    def _create_sensitivity_sheet(self, ws: Any, data: dict[str, Any]) -> None:
        """Create sensitivity analysis sheet in Excel."""
        from openpyxl.styles import Font, PatternFill

        ws["A1"] = "Sensitivity Analysis"
        ws["A1"].font = Font(bold=True, size=12)

        sensitivity = data.get("sensitivity", [])
        if not sensitivity:
            return

        waccs = sorted(set(item[0] for item in sensitivity))
        tgs = sorted(set(item[1] for item in sensitivity))

        # Headers
        ws["A3"] = "WACC \\ TG"
        for j, tg in enumerate(tgs):
            col = chr(ord("B") + j)
            ws[f"{col}3"] = f"{tg:.1%}"

        # Data
        for i, wacc in enumerate(waccs):
            row = 4 + i
            ws[f"A{row}"] = f"{wacc:.1%}"
            for j, tg in enumerate(tgs):
                col = chr(ord("B") + j)
                value = next(
                    (item[2] for item in sensitivity if item[0] == wacc and item[1] == tg),
                    0
                )
                ws[f"{col}{row}"] = value

    def _create_reverse_dcf_sheet(self, ws: Any, result: ReverseDCFResult) -> None:
        """Create reverse DCF sheet in Excel."""
        from openpyxl.styles import Font

        ws["A1"] = "Reverse DCF Analysis"
        ws["A1"].font = Font(bold=True, size=12)

        ws["A3"] = "Implied Revenue CAGR"
        ws["B3"] = result.implied_revenue_cagr

        ws["A4"] = "Market Cap"
        ws["B4"] = result.market_cap

        ws["A5"] = "Enterprise Value"
        ws["B5"] = result.enterprise_value

        ws["A6"] = "Implied Year 5 Revenue"
        ws["B6"] = result.implied_year5_revenue

        ws["A7"] = "Implied Year 5 FCF"
        ws["B7"] = result.implied_year5_fcf

        ws["A9"] = "Reasonableness"
        ws["B9"] = "Reasonable" if result.is_reasonable else "Questionable"

        ws["A11"] = "Notes:"
        for i, note in enumerate(result.reasonableness_notes):
            ws[f"A{12 + i}"] = note

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20
