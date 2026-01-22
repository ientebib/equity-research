"""
Professional DCF Valuation Model for GOOGL (Alphabet Inc.)
Built with openpyxl following investment banking standards
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Color standards for investment banking
BLUE_FONT = Font(color="0000FF", bold=False)  # Inputs/Assumptions
BLACK_FONT = Font(color="000000", bold=False)  # Formulas
GREEN_FONT = Font(color="008000", bold=False)  # Links from other sheets
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)

# Border styles
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_dcf_model():
    """Create a professional DCF valuation model for GOOGL"""

    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Create all sheets
    create_assumptions_sheet(wb)
    create_income_statement_sheet(wb)
    create_fcf_sheet(wb)
    create_dcf_valuation_sheet(wb)
    create_sensitivity_sheet(wb)

    # Save the workbook
    output_path = '/Users/isaacentebi/Desktop/Projects/equity-research/output/GOOGL_agent_sdk_test/GOOGL_DCF.xlsx'
    wb.save(output_path)
    print(f"DCF Model created successfully: {output_path}")
    return output_path


def create_assumptions_sheet(wb):
    """Sheet 1: Assumptions - All model inputs"""
    ws = wb.create_sheet("Assumptions", 0)

    # Title
    ws['A1'] = 'GOOGL (Alphabet Inc.) - DCF Model Assumptions'
    ws['A1'].font = TITLE_FONT
    ws['A2'] = f'As of: {datetime.now().strftime("%B %d, %Y")}'

    # Company Overview Section
    row = 4
    ws[f'A{row}'] = 'COMPANY OVERVIEW'
    ws[f'A{row}'].font = HEADER_FONT

    row += 1
    ws[f'A{row}'] = 'Company Name'
    ws[f'B{row}'] = 'Alphabet Inc. (GOOGL)'
    ws[f'B{row}'].font = BLUE_FONT

    row += 1
    ws[f'A{row}'] = 'Sector'
    ws[f'B{row}'] = 'Technology'
    ws[f'B{row}'].font = BLUE_FONT

    row += 1
    ws[f'A{row}'] = 'Base Year'
    ws[f'B{row}'] = 2024
    ws[f'B{row}'].font = BLUE_FONT

    # Current Financial Metrics Section
    row += 2
    ws[f'A{row}'] = 'CURRENT FINANCIAL METRICS (2024)'
    ws[f'A{row}'].font = HEADER_FONT

    row += 1
    ws[f'A{row}'] = 'Revenue ($B)'
    ws[f'B{row}'] = 307.4
    ws[f'B{row}'].font = BLUE_FONT
    ws[f'B{row}'].number_format = '#,##0.0'

    row += 1
    ws[f'A{row}'] = 'Operating Income / EBIT ($B)'
    ws[f'B{row}'] = 123.2
    ws[f'B{row}'].font = BLUE_FONT
    ws[f'B{row}'].number_format = '#,##0.0'

    row += 1
    ws[f'A{row}'] = 'Current EBIT Margin %'
    ws[f'B{row}'] = '=B11/B10'
    ws[f'B{row}'].font = BLACK_FONT
    ws[f'B{row}'].number_format = '0.0%'

    row += 1
    ws[f'A{row}'] = 'Cash ($B)'
    ws[f'B{row}'] = 86.7
    ws[f'B{row}'].font = BLUE_FONT
    ws[f'B{row}'].number_format = '#,##0.0'

    row += 1
    ws[f'A{row}'] = 'Total Debt ($B)'
    ws[f'B{row}'] = 15.9
    ws[f'B{row}'].font = BLUE_FONT
    ws[f'B{row}'].number_format = '#,##0.0'

    row += 1
    ws[f'A{row}'] = 'Total Equity ($B)'
    ws[f'B{row}'] = 245.7
    ws[f'B{row}'].font = BLUE_FONT
    ws[f'B{row}'].number_format = '#,##0.0'

    row += 1
    ws[f'A{row}'] = 'Market Cap ($B)'
    ws[f'B{row}'] = 2100
    ws[f'B{row}'].font = BLUE_FONT
    ws[f'B{row}'].number_format = '#,##0.0'

    row += 1
    ws[f'A{row}'] = 'Current Stock Price ($)'
    ws[f'B{row}'] = 180
    ws[f'B{row}'].font = BLUE_FONT
    ws[f'B{row}'].number_format = '#,##0.00'

    row += 1
    ws[f'A{row}'] = 'Shares Outstanding (B)'
    ws[f'B{row}'] = '=B16/B17'
    ws[f'B{row}'].font = BLACK_FONT
    ws[f'B{row}'].number_format = '#,##0.00'

    # WACC Calculation Section
    row += 2
    ws[f'A{row}'] = 'WACC CALCULATION'
    ws[f'A{row}'].font = HEADER_FONT

    row += 1
    ws[f'A{row}'] = 'Risk-Free Rate (10Y Treasury)'
    ws[f'B{row}'] = 0.045
    ws[f'B{row}'].font = BLUE_FONT
    ws[f'B{row}'].number_format = '0.0%'

    row += 1
    ws[f'A{row}'] = 'Beta'
    ws[f'B{row}'] = 1.2
    ws[f'B{row}'].font = BLUE_FONT
    ws[f'B{row}'].number_format = '0.0'

    row += 1
    ws[f'A{row}'] = 'Equity Risk Premium'
    ws[f'B{row}'] = 0.055
    ws[f'B{row}'].font = BLUE_FONT
    ws[f'B{row}'].number_format = '0.0%'

    row += 1
    ws[f'A{row}'] = 'Market Risk Premium (Beta * ERP)'
    ws[f'B{row}'] = '=B22*B23'
    ws[f'B{row}'].font = BLACK_FONT
    ws[f'B{row}'].number_format = '0.0%'

    row += 1
    ws[f'A{row}'] = 'Cost of Equity (Re)'
    ws[f'B{row}'] = '=B21+B24'
    ws[f'B{row}'].font = BLACK_FONT
    ws[f'B{row}'].number_format = '0.0%'

    row += 1
    ws[f'A{row}'] = 'Cost of Debt (Rd)'
    ws[f'B{row}'] = 0.04
    ws[f'B{row}'].font = BLUE_FONT
    ws[f'B{row}'].number_format = '0.0%'

    row += 1
    ws[f'A{row}'] = 'Tax Rate'
    ws[f'B{row}'] = 0.21
    ws[f'B{row}'].font = BLUE_FONT
    ws[f'B{row}'].number_format = '0.0%'

    row += 1
    ws[f'A{row}'] = 'Total Debt + Equity ($B)'
    ws[f'B{row}'] = '=B14+B15'
    ws[f'B{row}'].font = BLACK_FONT
    ws[f'B{row}'].number_format = '#,##0.0'

    row += 1
    ws[f'A{row}'] = 'Debt / (Debt + Equity) Ratio'
    ws[f'B{row}'] = '=B14/B29'
    ws[f'B{row}'].font = BLACK_FONT
    ws[f'B{row}'].number_format = '0.0%'

    row += 1
    ws[f'A{row}'] = 'Equity / (Debt + Equity) Ratio'
    ws[f'B{row}'] = '=B15/B29'
    ws[f'B{row}'].font = BLACK_FONT
    ws[f'B{row}'].number_format = '0.0%'

    row += 1
    ws[f'A{row}'] = 'WACC'
    ws[f'B{row}'] = '=(B31*B25)+(B30*B26*(1-B27))'
    ws[f'B{row}'].font = BLACK_FONT
    ws[f'B{row}'].number_format = '0.0%'
    ws[f'B{row}'].fill = YELLOW_FILL

    # Revenue Growth Assumptions Section
    row += 2
    ws[f'A{row}'] = 'REVENUE GROWTH RATES'
    ws[f'A{row}'].font = HEADER_FONT

    row += 1
    ws[f'A{row}'] = 'Year 1 Growth Rate'
    ws[f'B{row}'] = 0.155
    ws[f'B{row}'].font = BLUE_FONT
    ws[f'B{row}'].number_format = '0.0%'

    row += 1
    ws[f'A{row}'] = 'Year 2 Growth Rate'
    ws[f'B{row}'] = 0.127
    ws[f'B{row}'].font = BLUE_FONT
    ws[f'B{row}'].number_format = '0.0%'

    row += 1
    ws[f'A{row}'] = 'Year 3 Growth Rate'
    ws[f'B{row}'] = 0.10
    ws[f'B{row}'].font = BLUE_FONT
    ws[f'B{row}'].number_format = '0.0%'

    row += 1
    ws[f'A{row}'] = 'Year 4 Growth Rate'
    ws[f'B{row}'] = 0.08
    ws[f'B{row}'].font = BLUE_FONT
    ws[f'B{row}'].number_format = '0.0%'

    row += 1
    ws[f'A{row}'] = 'Year 5 Growth Rate'
    ws[f'B{row}'] = 0.06
    ws[f'B{row}'].font = BLUE_FONT
    ws[f'B{row}'].number_format = '0.0%'

    # Operating Margin Assumptions Section
    row += 2
    ws[f'A{row}'] = 'EBIT MARGIN ASSUMPTIONS'
    ws[f'A{row}'].font = HEADER_FONT

    row += 1
    ws[f'A{row}'] = 'Year 1 EBIT Margin'
    ws[f'B{row}'] = 0.400
    ws[f'B{row}'].font = BLUE_FONT
    ws[f'B{row}'].number_format = '0.0%'

    row += 1
    ws[f'A{row}'] = 'Year 2 EBIT Margin'
    ws[f'B{row}'] = 0.405
    ws[f'B{row}'].font = BLUE_FONT
    ws[f'B{row}'].number_format = '0.0%'

    row += 1
    ws[f'A{row}'] = 'Year 3 EBIT Margin'
    ws[f'B{row}'] = 0.410
    ws[f'B{row}'].font = BLUE_FONT
    ws[f'B{row}'].number_format = '0.0%'

    row += 1
    ws[f'A{row}'] = 'Year 4 EBIT Margin'
    ws[f'B{row}'] = 0.415
    ws[f'B{row}'].font = BLUE_FONT
    ws[f'B{row}'].number_format = '0.0%'

    row += 1
    ws[f'A{row}'] = 'Year 5 EBIT Margin'
    ws[f'B{row}'] = 0.420
    ws[f'B{row}'].font = BLUE_FONT
    ws[f'B{row}'].number_format = '0.0%'

    # Cash Flow Assumptions Section
    row += 2
    ws[f'A{row}'] = 'CASH FLOW ASSUMPTIONS'
    ws[f'A{row}'].font = HEADER_FONT

    row += 1
    ws[f'A{row}'] = 'D&A as % of Revenue'
    ws[f'B{row}'] = 0.06
    ws[f'B{row}'].font = BLUE_FONT
    ws[f'B{row}'].number_format = '0.0%'

    row += 1
    ws[f'A{row}'] = 'CapEx as % of Revenue'
    ws[f'B{row}'] = 0.12
    ws[f'B{row}'].font = BLUE_FONT
    ws[f'B{row}'].number_format = '0.0%'

    row += 1
    ws[f'A{row}'] = 'NWC as % of Revenue'
    ws[f'B{row}'] = -0.05
    ws[f'B{row}'].font = BLUE_FONT
    ws[f'B{row}'].number_format = '0.0%'

    # Terminal Value Assumptions Section
    row += 2
    ws[f'A{row}'] = 'TERMINAL VALUE ASSUMPTIONS'
    ws[f'A{row}'].font = HEADER_FONT

    row += 1
    ws[f'A{row}'] = 'Terminal Growth Rate'
    ws[f'B{row}'] = 0.03
    ws[f'B{row}'].font = BLUE_FONT
    ws[f'B{row}'].number_format = '0.0%'
    ws[f'B{row}'].fill = YELLOW_FILL

    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20

    # Freeze panes
    ws.freeze_panes = 'A4'


def create_income_statement_sheet(wb):
    """Sheet 2: Income Statement Projections (5 years)"""
    ws = wb.create_sheet("Income Statement", 1)

    # Title
    ws['A1'] = 'GOOGL - Income Statement Projections'
    ws['A1'].font = TITLE_FONT
    ws['A2'] = '$ in Billions'

    # Headers
    row = 4
    ws[f'A{row}'] = 'Period'
    ws[f'A{row}'].font = HEADER_FONT

    # Year headers
    years = ['Base Year\n2024', '2025E', '2026E', '2027E', '2028E', '2029E']
    for i, year in enumerate(years):
        col = get_column_letter(i + 2)
        ws[f'{col}{row}'] = year
        ws[f'{col}{row}'].font = HEADER_FONT
        ws[f'{col}{row}'].alignment = Alignment(horizontal='center', wrap_text=True)

    # Revenue
    row += 1
    ws[f'A{row}'] = 'Revenue'
    ws[f'A{row}'].font = HEADER_FONT

    # Base year revenue
    ws[f'B{row}'] = '=Assumptions!B10'
    ws[f'B{row}'].font = GREEN_FONT
    ws[f'B{row}'].number_format = '#,##0.0'

    # Projected years revenue
    growth_cells = ['B34', 'B35', 'B36', 'B37', 'B38']
    for i, growth_cell in enumerate(growth_cells):
        col = get_column_letter(i + 3)
        prev_col = get_column_letter(i + 2)
        ws[f'{col}{row}'] = f'={prev_col}{row}*(1+Assumptions!{growth_cell})'
        ws[f'{col}{row}'].font = BLACK_FONT
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # EBIT
    row += 1
    ws[f'A{row}'] = 'EBIT (Operating Income)'
    ws[f'A{row}'].font = HEADER_FONT

    # Base year EBIT
    ws[f'B{row}'] = '=Assumptions!B11'
    ws[f'B{row}'].font = GREEN_FONT
    ws[f'B{row}'].number_format = '#,##0.0'

    # Projected years EBIT
    margin_cells = ['B41', 'B42', 'B43', 'B44', 'B45']
    for i, margin_cell in enumerate(margin_cells):
        col = get_column_letter(i + 3)
        ws[f'{col}{row}'] = f'={col}5*Assumptions!{margin_cell}'
        ws[f'{col}{row}'].font = BLACK_FONT
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # EBIT Margin %
    row += 1
    ws[f'A{row}'] = 'EBIT Margin %'

    for i in range(6):
        col = get_column_letter(i + 2)
        ws[f'{col}{row}'] = f'={col}6/{col}5'
        ws[f'{col}{row}'].font = BLACK_FONT
        ws[f'{col}{row}'].number_format = '0.0%'

    # Revenue Growth %
    row += 1
    ws[f'A{row}'] = 'Revenue Growth %'

    ws[f'B{row}'] = 'N/A'
    ws[f'B{row}'].font = BLACK_FONT

    for i in range(5):
        col = get_column_letter(i + 3)
        prev_col = get_column_letter(i + 2)
        ws[f'{col}{row}'] = f'=({col}5/{prev_col}5)-1'
        ws[f'{col}{row}'].font = BLACK_FONT
        ws[f'{col}{row}'].number_format = '0.0%'

    # Set column widths
    ws.column_dimensions['A'].width = 30
    for i in range(2, 8):
        ws.column_dimensions[get_column_letter(i)].width = 15

    # Freeze panes
    ws.freeze_panes = 'B5'


def create_fcf_sheet(wb):
    """Sheet 3: Free Cash Flow Build"""
    ws = wb.create_sheet("Free Cash Flow", 2)

    # Title
    ws['A1'] = 'GOOGL - Unlevered Free Cash Flow'
    ws['A1'].font = TITLE_FONT
    ws['A2'] = '$ in Billions'

    # Headers
    row = 4
    ws[f'A{row}'] = 'Period'
    ws[f'A{row}'].font = HEADER_FONT

    # Year headers (projection years only)
    years = ['2025E', '2026E', '2027E', '2028E', '2029E']
    for i, year in enumerate(years):
        col = get_column_letter(i + 2)
        ws[f'{col}{row}'] = year
        ws[f'{col}{row}'].font = HEADER_FONT
        ws[f'{col}{row}'].alignment = Alignment(horizontal='center')

    # EBIT
    row += 1
    ws[f'A{row}'] = 'EBIT'

    for i in range(5):
        col = get_column_letter(i + 2)
        income_col = get_column_letter(i + 3)
        ws[f'{col}{row}'] = f"='Income Statement'!{income_col}6"
        ws[f'{col}{row}'].font = GREEN_FONT
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Taxes
    row += 1
    ws[f'A{row}'] = 'Less: Taxes @ 21%'

    for i in range(5):
        col = get_column_letter(i + 2)
        ws[f'{col}{row}'] = f'={col}5*Assumptions!B27'
        ws[f'{col}{row}'].font = BLACK_FONT
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # NOPAT
    row += 1
    ws[f'A{row}'] = 'NOPAT (EBIT after Tax)'
    ws[f'A{row}'].font = HEADER_FONT

    for i in range(5):
        col = get_column_letter(i + 2)
        ws[f'{col}{row}'] = f'={col}5-{col}6'
        ws[f'{col}{row}'].font = BLACK_FONT
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # D&A
    row += 1
    ws[f'A{row}'] = 'Add: Depreciation & Amortization'

    for i in range(5):
        col = get_column_letter(i + 2)
        income_col = get_column_letter(i + 3)
        ws[f'{col}{row}'] = f"='Income Statement'!{income_col}5*Assumptions!B49"
        ws[f'{col}{row}'].font = BLACK_FONT
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # CapEx
    row += 1
    ws[f'A{row}'] = 'Less: Capital Expenditures'

    for i in range(5):
        col = get_column_letter(i + 2)
        income_col = get_column_letter(i + 3)
        ws[f'{col}{row}'] = f"='Income Statement'!{income_col}5*Assumptions!B50"
        ws[f'{col}{row}'].font = BLACK_FONT
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # NWC (Net Working Capital)
    row += 1
    ws[f'A{row}'] = 'Net Working Capital'

    # Base year NWC (Year 0)
    ws['A11'] = 'Base Year NWC (2024)'
    ws['B11'] = '=Assumptions!B10*Assumptions!B51'
    ws['B11'].font = BLACK_FONT
    ws['B11'].number_format = '#,##0.0'

    # Projected NWC for each year
    for i in range(5):
        col = get_column_letter(i + 2)
        income_col = get_column_letter(i + 3)
        ws[f'{col}{row}'] = f"='Income Statement'!{income_col}5*Assumptions!B51"
        ws[f'{col}{row}'].font = BLACK_FONT
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Change in NWC
    row += 1
    ws[f'A{row}'] = 'Less: Change in NWC'

    for i in range(5):
        col = get_column_letter(i + 2)
        prev_col = get_column_letter(i + 1) if i > 0 else 'B'
        prev_row = 10 if i > 0 else 11
        ws[f'{col}{row}'] = f'={col}10-{prev_col}{prev_row}'
        ws[f'{col}{row}'].font = BLACK_FONT
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Unlevered Free Cash Flow
    row += 1
    ws[f'A{row}'] = 'Unlevered Free Cash Flow'
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = YELLOW_FILL

    for i in range(5):
        col = get_column_letter(i + 2)
        ws[f'{col}{row}'] = f'={col}7+{col}8-{col}9-{col}11'
        ws[f'{col}{row}'].font = BLACK_FONT
        ws[f'{col}{row}'].number_format = '#,##0.0'
        ws[f'{col}{row}'].fill = YELLOW_FILL

    # Set column widths
    ws.column_dimensions['A'].width = 35
    for i in range(2, 7):
        ws.column_dimensions[get_column_letter(i)].width = 15

    # Freeze panes
    ws.freeze_panes = 'B5'


def create_dcf_valuation_sheet(wb):
    """Sheet 4: DCF Valuation"""
    ws = wb.create_sheet("DCF Valuation", 3)

    # Title
    ws['A1'] = 'GOOGL - DCF Valuation Analysis'
    ws['A1'].font = TITLE_FONT
    ws['A2'] = '$ in Billions (except per share data)'

    # Headers
    row = 4
    ws[f'A{row}'] = 'Period'
    ws[f'A{row}'].font = HEADER_FONT

    # Year headers
    years = ['2025E', '2026E', '2027E', '2028E', '2029E']
    for i, year in enumerate(years):
        col = get_column_letter(i + 2)
        ws[f'{col}{row}'] = year
        ws[f'{col}{row}'].font = HEADER_FONT
        ws[f'{col}{row}'].alignment = Alignment(horizontal='center')

    # Unlevered Free Cash Flow
    row += 1
    ws[f'A{row}'] = 'Unlevered Free Cash Flow'

    for i in range(5):
        col = get_column_letter(i + 2)
        ws[f'{col}{row}'] = f"='Free Cash Flow'!{col}12"
        ws[f'{col}{row}'].font = GREEN_FONT
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Discount Period
    row += 1
    ws[f'A{row}'] = 'Discount Period (Years)'

    for i in range(5):
        col = get_column_letter(i + 2)
        ws[f'{col}{row}'] = i + 1
        ws[f'{col}{row}'].font = BLUE_FONT
        ws[f'{col}{row}'].number_format = '0'

    # WACC
    row += 1
    ws[f'A{row}'] = 'WACC'

    for i in range(5):
        col = get_column_letter(i + 2)
        ws[f'{col}{row}'] = '=Assumptions!B32'
        ws[f'{col}{row}'].font = GREEN_FONT
        ws[f'{col}{row}'].number_format = '0.0%'

    # Discount Factor
    row += 1
    ws[f'A{row}'] = 'Discount Factor'

    for i in range(5):
        col = get_column_letter(i + 2)
        ws[f'{col}{row}'] = f'=1/((1+{col}7)^{col}6)'
        ws[f'{col}{row}'].font = BLACK_FONT
        ws[f'{col}{row}'].number_format = '0.000'

    # Present Value of FCF
    row += 1
    ws[f'A{row}'] = 'PV of FCF'

    for i in range(5):
        col = get_column_letter(i + 2)
        ws[f'{col}{row}'] = f'={col}5*{col}8'
        ws[f'{col}{row}'].font = BLACK_FONT
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Sum of PV of FCFs
    row += 2
    ws[f'A{row}'] = 'Sum of PV of FCFs (2025E-2029E)'
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'B{row}'] = '=SUM(B9:F9)'
    ws[f'B{row}'].font = BLACK_FONT
    ws[f'B{row}'].number_format = '#,##0.0'

    # Terminal Value Calculation
    row += 2
    ws[f'A{row}'] = 'TERMINAL VALUE CALCULATION'
    ws[f'A{row}'].font = TITLE_FONT

    row += 1
    ws[f'A{row}'] = 'Year 5 FCF (2029E)'
    ws[f'B{row}'] = "='Free Cash Flow'!F12"
    ws[f'B{row}'].font = GREEN_FONT
    ws[f'B{row}'].number_format = '#,##0.0'

    row += 1
    ws[f'A{row}'] = 'Terminal Growth Rate'
    ws[f'B{row}'] = '=Assumptions!B54'
    ws[f'B{row}'].font = GREEN_FONT
    ws[f'B{row}'].number_format = '0.0%'

    row += 1
    ws[f'A{row}'] = 'WACC'
    ws[f'B{row}'] = '=Assumptions!B32'
    ws[f'B{row}'].font = GREEN_FONT
    ws[f'B{row}'].number_format = '0.0%'

    row += 1
    ws[f'A{row}'] = 'Terminal Value'
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'B{row}'] = '=(B14*(1+B15))/(B16-B15)'
    ws[f'B{row}'].font = BLACK_FONT
    ws[f'B{row}'].number_format = '#,##0.0'

    row += 1
    ws[f'A{row}'] = 'Discount Factor (Year 5)'
    ws[f'B{row}'] = '=F8'
    ws[f'B{row}'].font = GREEN_FONT
    ws[f'B{row}'].number_format = '0.000'

    row += 1
    ws[f'A{row}'] = 'PV of Terminal Value'
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'B{row}'] = '=B17*B18'
    ws[f'B{row}'].font = BLACK_FONT
    ws[f'B{row}'].number_format = '#,##0.0'

    # Enterprise Value
    row += 2
    ws[f'A{row}'] = 'ENTERPRISE VALUE'
    ws[f'A{row}'].font = TITLE_FONT

    row += 1
    ws[f'A{row}'] = 'PV of FCFs (2025E-2029E)'
    ws[f'B{row}'] = '=B11'
    ws[f'B{row}'].font = GREEN_FONT
    ws[f'B{row}'].number_format = '#,##0.0'

    row += 1
    ws[f'A{row}'] = 'PV of Terminal Value'
    ws[f'B{row}'] = '=B19'
    ws[f'B{row}'].font = GREEN_FONT
    ws[f'B{row}'].number_format = '#,##0.0'

    row += 1
    ws[f'A{row}'] = 'Enterprise Value'
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'B{row}'] = '=B22+B23'
    ws[f'B{row}'].font = BLACK_FONT
    ws[f'B{row}'].number_format = '#,##0.0'
    ws[f'B{row}'].fill = YELLOW_FILL

    # Equity Value
    row += 2
    ws[f'A{row}'] = 'EQUITY VALUE CALCULATION'
    ws[f'A{row}'].font = TITLE_FONT

    row += 1
    ws[f'A{row}'] = 'Enterprise Value'
    ws[f'B{row}'] = '=B24'
    ws[f'B{row}'].font = GREEN_FONT
    ws[f'B{row}'].number_format = '#,##0.0'

    row += 1
    ws[f'A{row}'] = 'Add: Cash'
    ws[f'B{row}'] = '=Assumptions!B13'
    ws[f'B{row}'].font = GREEN_FONT
    ws[f'B{row}'].number_format = '#,##0.0'

    row += 1
    ws[f'A{row}'] = 'Less: Total Debt'
    ws[f'B{row}'] = '=Assumptions!B14'
    ws[f'B{row}'].font = GREEN_FONT
    ws[f'B{row}'].number_format = '#,##0.0'

    row += 1
    ws[f'A{row}'] = 'Equity Value'
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'B{row}'] = '=B27+B28-B29'
    ws[f'B{row}'].font = BLACK_FONT
    ws[f'B{row}'].number_format = '#,##0.0'
    ws[f'B{row}'].fill = YELLOW_FILL

    # Per Share Valuation
    row += 2
    ws[f'A{row}'] = 'PER SHARE VALUATION'
    ws[f'A{row}'].font = TITLE_FONT

    row += 1
    ws[f'A{row}'] = 'Equity Value ($B)'
    ws[f'B{row}'] = '=B30'
    ws[f'B{row}'].font = GREEN_FONT
    ws[f'B{row}'].number_format = '#,##0.0'

    row += 1
    ws[f'A{row}'] = 'Shares Outstanding (B)'
    ws[f'B{row}'] = '=Assumptions!B18'
    ws[f'B{row}'].font = GREEN_FONT
    ws[f'B{row}'].number_format = '#,##0.00'

    row += 1
    ws[f'A{row}'] = 'Implied Price Per Share'
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'B{row}'] = '=B33/B34'
    ws[f'B{row}'].font = BLACK_FONT
    ws[f'B{row}'].number_format = '$#,##0.00'
    ws[f'B{row}'].fill = YELLOW_FILL

    row += 1
    ws[f'A{row}'] = 'Current Stock Price'
    ws[f'B{row}'] = '=Assumptions!B17'
    ws[f'B{row}'].font = GREEN_FONT
    ws[f'B{row}'].number_format = '$#,##0.00'

    row += 1
    ws[f'A{row}'] = 'Upside/(Downside) %'
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'B{row}'] = '=(B35/B36)-1'
    ws[f'B{row}'].font = BLACK_FONT
    ws[f'B{row}'].number_format = '0.0%'
    ws[f'B{row}'].fill = YELLOW_FILL

    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    for i in range(3, 7):
        ws.column_dimensions[get_column_letter(i)].width = 15

    # Freeze panes
    ws.freeze_panes = 'B5'


def create_sensitivity_sheet(wb):
    """Sheet 5: Sensitivity Analysis"""
    ws = wb.create_sheet("Sensitivity Analysis", 4)

    # Title
    ws['A1'] = 'GOOGL - DCF Sensitivity Analysis'
    ws['A1'].font = TITLE_FONT
    ws['A2'] = 'Implied Share Price Sensitivity to WACC and Terminal Growth Rate'

    # Create sensitivity table
    row = 4
    ws[f'A{row}'] = 'WACC \\ Terminal Growth'
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].alignment = Alignment(horizontal='center', wrap_text=True)

    # Terminal growth rate headers (columns)
    terminal_growth_rates = [0.020, 0.025, 0.030, 0.035, 0.040]
    for i, rate in enumerate(terminal_growth_rates):
        col = get_column_letter(i + 2)
        ws[f'{col}{row}'] = rate
        ws[f'{col}{row}'].font = HEADER_FONT
        ws[f'{col}{row}'].number_format = '0.0%'
        ws[f'{col}{row}'].alignment = Alignment(horizontal='center')

    # WACC rates (rows) and sensitivity values
    wacc_rates = [0.060, 0.070, 0.080, 0.090, 0.100]

    for i, wacc in enumerate(wacc_rates):
        row += 1
        # WACC label
        ws[f'A{row}'] = wacc
        ws[f'A{row}'].font = HEADER_FONT
        ws[f'A{row}'].number_format = '0.0%'
        ws[f'A{row}'].alignment = Alignment(horizontal='center')

        # Sensitivity values
        for j, term_growth in enumerate(terminal_growth_rates):
            col = get_column_letter(j + 2)

            # Formula to calculate implied share price with different WACC and terminal growth
            # We need to recalculate the entire DCF with new assumptions
            formula = f'''
=(
    ('Free Cash Flow'!B12/(1+A{row})^1) +
    ('Free Cash Flow'!C12/(1+A{row})^2) +
    ('Free Cash Flow'!D12/(1+A{row})^3) +
    ('Free Cash Flow'!E12/(1+A{row})^4) +
    ('Free Cash Flow'!F12/(1+A{row})^5) +
    ((('Free Cash Flow'!F12*(1+{col}$4))/(A{row}-{col}$4))/(1+A{row})^5)
    + Assumptions!B13 - Assumptions!B14
) / Assumptions!B18
'''.replace('\n', '').replace('    ', '')

            ws[f'{col}{row}'] = formula
            ws[f'{col}{row}'].font = BLACK_FONT
            ws[f'{col}{row}'].number_format = '$#,##0.00'

            # Highlight base case (8.0% WACC, 3.0% terminal growth)
            if wacc == 0.080 and term_growth == 0.030:
                ws[f'{col}{row}'].fill = YELLOW_FILL

    # Add reference to base case
    row += 2
    ws[f'A{row}'] = 'Base Case Highlighted (8.0% WACC, 3.0% Terminal Growth)'
    ws[f'A{row}'].font = Font(italic=True)

    # Add note
    row += 1
    ws[f'A{row}'] = 'Note: Sensitivity analysis shows implied share price under different WACC and terminal growth scenarios'
    ws[f'A{row}'].font = Font(italic=True, size=9)

    # Set column widths
    ws.column_dimensions['A'].width = 15
    for i in range(2, 7):
        ws.column_dimensions[get_column_letter(i)].width = 15

    # Freeze panes
    ws.freeze_panes = 'B5'


if __name__ == "__main__":
    create_dcf_model()
